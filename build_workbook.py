"""
Build an Excel workbook from the analysis output.

A PR team works in spreadsheets, not CSVs. This produces one file with the
ranking, the segment breakdowns and the robustness check on separate tabs,
formatted so it can be read without any further work.

    pip3 install openpyxl
    python3 build_workbook.py

Output: reboot_output/Scotland_picturesque_routes.xlsx
"""

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent / "reboot_output"
PROCESSED = ROOT / "data_processed"
OUT = ROOT / "Scotland_picturesque_routes.xlsx"

INK = "1D3557"
ACCENT = "E07A5F"
BAND = "F6F8FA"

HEAD = Font(color="FFFFFF", bold=True, size=10.5)
HEAD_FILL = PatternFill("solid", fgColor=INK)
TITLE = Font(color=INK, bold=True, size=14)
NOTE = Font(color="5B6670", italic=True, size=9)
THIN = Side(style="thin", color="DDE3E8")
BORDER = Border(bottom=THIN)


def write_sheet(writer, df, sheet, title, note="", widths=None,
                pct_cols=(), num_cols=()):
    """Write one dataframe with a title row, styled header and banding."""
    df.to_excel(writer, sheet_name=sheet, startrow=3, index=False)
    ws = writer.sheets[sheet]

    ws["A1"] = title
    ws["A1"].font = TITLE
    if note:
        ws["A2"] = note
        ws["A2"].font = NOTE

    ncols = len(df.columns)

    for c in range(1, ncols + 1):
        cell = ws.cell(row=4, column=c)
        cell.font = HEAD
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 30

    for r in range(5, len(df) + 5):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if r % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=BAND)
            header = df.columns[c - 1]
            if header in pct_cols:
                cell.number_format = "0.0%"
            elif header in num_cols:
                cell.number_format = "#,##0.00"

    for i, col in enumerate(df.columns, start=1):
        letter = get_column_letter(i)
        if widths and col in widths:
            ws.column_dimensions[letter].width = widths[col]
        else:
            longest = max([len(str(col))] +
                          [len(str(v)) for v in df[col].head(60)])
            ws.column_dimensions[letter].width = min(max(longest + 2, 10), 46)

    ws.freeze_panes = "A5"


def main() -> None:
    scored = pd.read_csv(PROCESSED / "routes_scored.csv")

    ranking = (scored[scored["headline_eligible"]]
               .sort_values("picturesque_score", ascending=False)
               .head(25)
               .loc[:, ["name", "picturesque_score", "avg_rating",
                        "num_reviews", "num_photos", "photos_per_review",
                        "scenery_mention_rate", "length_km",
                        "elevation_gain_m", "difficulty", "url"]]
               .reset_index(drop=True))
    ranking.insert(0, "rank", range(1, len(ranking) + 1))
    ranking.columns = ["Rank", "Route", "Picturesque index", "Rating",
                       "Reviews", "Photos", "Photos per review",
                       "Scenery mentions", "Length (km)", "Ascent (m)",
                       "Difficulty", "AllTrails URL"]

    full = (scored.sort_values("picturesque_score", ascending=False)
            .loc[:, ["name", "picturesque_score", "headline_eligible",
                     "avg_rating", "num_reviews", "num_photos",
                     "photos_per_review", "scenery_mention_rate",
                     "detractor_rate", "length_km", "elevation_gain_m",
                     "elevation_per_km", "difficulty", "route_type",
                     "latitude", "longitude", "components_used", "url"]]
            .reset_index(drop=True))
    full.columns = ["Route", "Picturesque index", "Eligible for ranking",
                    "Rating", "Reviews", "Photos", "Photos per review",
                    "Scenery mentions", "Detractor mentions", "Length (km)",
                    "Ascent (m)", "Ascent per km", "Difficulty", "Route type",
                    "Latitude", "Longitude", "Signals used", "AllTrails URL"]

    by_diff = pd.read_csv(PROCESSED / "summary_by_difficulty.csv")
    by_diff.columns = ["Difficulty", "Routes", "Mean index", "Mean rating",
                       "Mean photos per review", "Mean length (km)"]

    sens = pd.read_csv(PROCESSED / "summary_sensitivity.csv")
    sens.columns = ["Weighting variant", "Rank correlation with chosen",
                    "Top-10 routes retained"]

    tags = pd.read_csv(PROCESSED / "summary_tag_counts.csv")
    tags.columns = ["AllTrails tag", "Routes carrying it"]

    corr = pd.read_csv(PROCESSED / "summary_correlations.csv", index_col=0)
    corr = corr.reset_index().rename(columns={"index": "Signal"})

    method = pd.DataFrame({
        "Signal": ["Photos per review", "Scenery language in reviews",
                   "AllTrails view tag", "Adjusted star rating",
                   "Elevation per km"],
        "Weight": [0.25, 0.25, 0.20, 0.15, 0.15],
        "What it measures": [
            "Did people stop mid-run to photograph it?",
            "Share of reviews using scenery language",
            "AllTrails' own three-level view rating",
            "Rating, shrunk toward the sample mean by review count",
            "Steepness as a proxy for viewpoints"],
    })

    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        write_sheet(writer, ranking, "Top 25",
                    "Scotland's most picturesque running routes",
                    "Routes with at least 10 reviews. Collected from AllTrails, "
                    "31 July 2026.",
                    widths={"Route": 46, "AllTrails URL": 40},
                    pct_cols=["Scenery mentions"],
                    num_cols=["Picturesque index", "Photos per review",
                              "Length (km)"])

        write_sheet(writer, method, "Index method",
                    "How the picturesque index is built",
                    "Each signal is converted to a percentile rank across the "
                    "sample, then weighted. Weights sum to 1.",
                    widths={"What it measures": 52}, pct_cols=["Weight"])

        write_sheet(writer, by_diff, "By difficulty",
                    "Photography and enjoyment by difficulty band",
                    "Ratings are flat across bands; photos per visitor more "
                    "than double from easy to hard.",
                    num_cols=["Mean index", "Mean rating",
                              "Mean photos per review", "Mean length (km)"])

        write_sheet(writer, sens, "Robustness",
                    "Does the ranking survive different weightings?",
                    "Index recomputed under equal weights and five "
                    "leave-one-out variants. Broad ordering is stable; exact "
                    "top-10 membership is not.",
                    widths={"Weighting variant": 34})

        write_sheet(writer, corr, "Correlations",
                    "Spearman correlation between signals",
                    "Rating vs photos per review is 0.13 - enjoyment and "
                    "photography measure different things.",
                    widths={"Signal": 24})

        write_sheet(writer, tags, "Tag counts",
                    "AllTrails tags across the sample",
                    "Tags are condition and amenity labels aggregated from "
                    "reviews, not terrain features.",
                    widths={"AllTrails tag": 26})

        write_sheet(writer, full, "Full dataset",
                    "All 100 collected routes, scored",
                    "Includes routes below the 10-review threshold, flagged in "
                    "the eligibility column.",
                    widths={"Route": 46, "AllTrails URL": 40},
                    pct_cols=["Scenery mentions", "Detractor mentions"],
                    num_cols=["Picturesque index", "Photos per review",
                              "Length (km)", "Ascent per km"])

    print(f"Written: {OUT}")
    print("Sheets: Top 25, Index method, By difficulty, Robustness, "
          "Correlations, Tag counts, Full dataset")


if __name__ == "__main__":
    main()
